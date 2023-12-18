# -*- coding: utf-8 -*-
from urllib.parse import quote

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework.decorators import action
from rest_framework.request import Request

from dvadmin.utils.import_export import import_to_data
from dvadmin.utils.json_response import DetailResponse
from dvadmin.utils.request_util import get_verbose_name


class ImportSerializerMixin:
    """
    Customize import templates and import functions
    """

    #
    import_field_dict = {}
    #
    import_serializer_class = None
    # Maximum width of table header，default 50
    export_column_width = 50

    def is_number(self, num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata

            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return (
            round(length, 1)
            if length <= self.export_column_width
            else self.export_column_width
        )

    @action(methods=["get", "post"], detail=False)
    @transaction.atomic
    def import_data(self, request: Request, *args, **kwargs):
        """

        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        assert self.import_field_dict, (
            "'%s', please configure the corresponding export template fields"
            % self.__class__.__name__
        )
        # export template
        if request.method == "GET":
            #
            queryset = self.filter_queryset(self.get_queryset())
            # export excel table
            response = HttpResponse(content_type="application/msexcel")
            response["Access-Control-Expose-Headers"] = f"Content-Disposition"
            response[
                "Content-Disposition"
            ] = f'attachment;filename={quote(str(f"export_{get_verbose_name(queryset)}_template.xlsx"))}'
            wb = Workbook()
            ws1 = wb.create_sheet("data", 1)
            ws1.sheet_state = "hidden"
            ws = wb.active
            row = get_column_letter(len(self.import_field_dict) + 1)
            column = 10
            header_data = [
                "ID",
            ]
            validation_data_dict = {}
            for index, ele in enumerate(self.import_field_dict.values()):
                if isinstance(ele, dict):
                    header_data.append(ele.get("title"))
                    choices = ele.get("choices", {})
                    if choices.get("data"):
                        data_list = []
                        data_list.extend(choices.get("data").keys())
                        validation_data_dict[ele.get("title")] = data_list
                    elif choices.get("queryset") and choices.get("values_name"):
                        data_list = choices.get("queryset").values_list(
                            choices.get("values_name"), flat=True
                        )
                        validation_data_dict[ele.get("title")] = list(data_list)
                    else:
                        continue
                    column_letter = get_column_letter(len(validation_data_dict))
                    dv = DataValidation(
                        type="list",
                        formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[ele.get('title')]) + 1}",
                        allow_blank=True,
                    )
                    ws.add_data_validation(dv)
                    dv.add(
                        f"{get_column_letter(index + 2)}2:{get_column_letter(index + 2)}1048576"
                    )
                else:
                    header_data.append(ele)
            #
            ws1.append(list(validation_data_dict.keys()))
            for index, validation_data in enumerate(validation_data_dict.values()):
                for inx, ele in enumerate(validation_data):
                    ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
            #
            df_len_max = [self.get_string_len(ele) for ele in header_data]
            ws.append(header_data)
            # 　update column width
            for index, width in enumerate(df_len_max):
                ws.column_dimensions[get_column_letter(index + 1)].width = width
            tab = Table(displayName="Table1", ref=f"A1:{row}{column}")
            style = TableStyleInfo(
                name="TableStyleLight11",
                showFirstColumn=True,
                showLastColumn=True,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            wb.save(response)
            return response

        updateSupport = request.data.get("updateSupport")
        # Organize the corresponding data structures from Excel and save them using a serializer
        queryset = self.filter_queryset(self.get_queryset())
        # get many-to-many fields
        m2m_fields = [
            ele.name
            for ele in queryset.model._meta.get_fields()
            if hasattr(ele, "many_to_many") and ele.many_to_many is True
        ]
        data = import_to_data(
            request.data.get("url"), self.import_field_dict, m2m_fields
        )
        unique_list = [
            ele.name
            for ele in queryset.model._meta.get_fields()
            if hasattr(ele, "unique") and ele.unique is True
        ]
        for ele in data:
            # get unique field
            if queryset.model._meta.unique_together:
                filter_dic = {
                    i: ele.get(i) for i in list(queryset.model._meta.unique_together[0])
                }
            else:
                filter_dic = {
                    i: ele.get(i)
                    for i in list(set(unique_list))
                    if ele.get(i) is not None
                }
            instance = filter_dic and queryset.filter(**filter_dic).first()
            if instance and not updateSupport:
                continue
            if not filter_dic:
                instance = None
            serializer = self.import_serializer_class(
                instance, data=ele, request=request
            )
            serializer.is_valid(raise_exception=True)
            serializer.save()
        return DetailResponse(msg=f"import succeed！")


class ExportSerializerMixin:
    """
    custom export serializer
    """

    #
    export_field_label = []
    #
    export_serializer_class = None
    #
    export_column_width = 50

    def is_number(self, num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata

            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return (
            round(length, 1)
            if length <= self.export_column_width
            else self.export_column_width
        )

    def export_data(self, request: Request, *args, **kwargs):
        """

        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        queryset = self.filter_queryset(self.get_queryset())
        assert self.export_field_label, (
            "'%s', please configure the corresponding export template fields"
            % self.__class__.__name__
        )
        assert self.export_serializer_class, (
            "'%s', please configure the corresponding export serializer"
            % self.__class__.__name__
        )
        data = self.export_serializer_class(queryset, many=True, request=request).data
        # export excel table
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response[
            "content-disposition"
        ] = f'attachment;filename={quote(str(f"export_{get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws = wb.active
        header_data = ["ID", *self.export_field_label.values()]
        hidden_header = ["#", *self.export_field_label.keys()]
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(self.export_field_label) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key, val in results.items():
                    if key == h_item:
                        if val is None or val == "":
                            results_list.append("")
                        else:
                            results_list.append(val)
                        #
                        result_column_width = self.get_string_len(val)
                        if h_index != 0 and result_column_width > df_len_max[h_index]:
                            df_len_max[h_index] = result_column_width
            ws.append([index + 1, *results_list])
            column += 1
        #
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  #
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response
